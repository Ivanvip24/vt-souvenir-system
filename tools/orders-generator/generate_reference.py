#!/usr/bin/env python3
"""
PDF Reference Sheet Generator
Automatically creates reference sheets for laser-cut souvenir orders
"""

import os
import sys
from datetime import datetime
from pathlib import Path

from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfform
from reportlab.platypus import Table, TableStyle
from PIL import Image as PILImage
from openpyxl import load_workbook
import yaml


class ReferenceSheetGenerator:
    """Generates PDF reference sheets from order data"""

    def __init__(self, config_path='config.yaml'):
        """Initialize generator with configuration"""
        self.load_config(config_path)

    def load_config(self, config_path):
        """Load configuration from YAML file"""
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)

        # Set up page dimensions
        if self.config['layout']['page_size'] == 'A4':
            self.page_width, self.page_height = A4
        else:
            self.page_width, self.page_height = letter

        self.margin = self.config['layout']['margin'] * mm
        self.columns = self.config['layout']['columns']

    def read_order_from_excel(self, excel_path):
        """
        Read order data from Excel file
        Expected columns: Product Name, Type, Quantity, Image Path
        """
        wb = load_workbook(excel_path)
        sheet = wb.active

        order_data = {
            'order_number': '',
            'client_name': '',
            'items': []
        }

        # Read header row to get order info
        if sheet.cell(1, 1).value == 'Order Number':
            order_data['order_number'] = str(sheet.cell(1, 2).value or '')
        if sheet.cell(2, 1).value == 'Client Name':
            order_data['client_name'] = str(sheet.cell(2, 2).value or '')

        # Read items starting from row 5 (skip header)
        header_row = 4
        for row in range(header_row + 1, sheet.max_row + 1):
            product_name = sheet.cell(row, 1).value
            product_type = sheet.cell(row, 2).value
            quantity = sheet.cell(row, 3).value
            image_path = sheet.cell(row, 4).value

            if not product_name:  # Skip empty rows
                continue

            order_data['items'].append({
                'name': str(product_name),
                'type': str(product_type),
                'quantity': int(quantity) if quantity else 0,
                'image_path': str(image_path) if image_path else None
            })

        return order_data

    def generate_pdf(self, order_data, output_filename=None):
        """Generate PDF reference sheet from order data"""

        if not output_filename:
            order_num = order_data.get('order_number', 'TEMP')
            output_filename = f"ORDER_{order_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        # Ensure output directory exists
        output_dir = Path(self.config['output_path'])
        output_dir.mkdir(parents=True, exist_ok=True)

        output_path = output_dir / output_filename

        # Create PDF
        c = canvas.Canvas(str(output_path), pagesize=(self.page_width, self.page_height))

        # Draw header
        self._draw_header(c, order_data)

        # Draw grid with items
        self._draw_items_grid(c, order_data['items'])

        # Save PDF
        c.save()

        print(f"[SUCCESS] PDF generated successfully: {output_path}")
        return str(output_path)

    def _draw_header(self, c, order_data):
        """Draw PDF header with order information"""
        c.setFont("Helvetica-Bold", self.config['fonts']['title_size'])
        c.drawString(self.margin, self.page_height - self.margin, "REFERENCE SHEET")

        # Order info
        c.setFont("Helvetica", self.config['fonts']['body_size'])
        y_pos = self.page_height - self.margin - 25

        order_info = f"Order: {order_data.get('order_number', 'N/A')}"
        if order_data.get('client_name'):
            order_info += f" | Client: {order_data['client_name']}"
        order_info += f" | Date: {datetime.now().strftime('%Y-%m-%d')}"

        c.drawString(self.margin, y_pos, order_info)

        # Draw line
        c.setLineWidth(2)
        c.line(self.margin, y_pos - 10, self.page_width - self.margin, y_pos - 10)

    def _draw_items_grid(self, c, items):
        """Draw grid of items"""
        cell_width = self.config['cell']['width']
        cell_height = self.config['cell']['height']
        padding = self.config['cell']['padding']

        # Calculate starting position
        start_y = self.page_height - self.margin - 60
        start_x = self.margin

        # Calculate grid layout
        rows_per_page = int((self.page_height - 2 * self.margin - 80) / cell_height)

        current_x = start_x
        current_y = start_y
        items_in_row = 0

        for idx, item in enumerate(items):
            # Check if we need a new page
            if current_y < cell_height + self.margin:
                c.showPage()
                current_y = self.page_height - self.margin
                current_x = start_x
                items_in_row = 0

            # Draw cell border
            c.setStrokeColor(colors.black)
            c.setLineWidth(1)
            c.rect(current_x, current_y - cell_height, cell_width, cell_height)

            # Draw item content
            self._draw_item_cell(c, item, current_x, current_y, cell_width, cell_height, padding)

            # Move to next position
            items_in_row += 1
            if items_in_row >= self.columns:
                # Move to next row
                current_y -= cell_height
                current_x = start_x
                items_in_row = 0
            else:
                # Move to next column
                current_x += cell_width

    def _draw_item_cell(self, c, item, x, y, width, height, padding):
        """Draw a single item cell"""
        # Type label at top
        c.setFont("Helvetica-Bold", self.config['fonts']['label_size'])
        c.drawString(x + padding, y - padding - 12, f"Tipo: {item['type']}")

        # Image in center
        if item.get('image_path') and os.path.exists(item['image_path']):
            self._draw_image(c, item['image_path'], x, y, width, height, padding)
        else:
            # No image placeholder
            c.setFont("Helvetica", self.config['fonts']['body_size'])
            text_y = y - height/2
            c.drawCentredString(x + width/2, text_y, item['name'])

        # Quantity ordered
        c.setFont("Helvetica-Bold", self.config['fonts']['body_size'])
        qty_y = y - height + padding + 60
        c.drawString(x + padding, qty_y, f"Requeridos: {item['quantity']}")

        # Editable fields
        if self.config['form_fields']['enable_editable']:
            field_height = self.config['form_fields']['field_height']
            field_y = qty_y - 25

            # Counted field
            c.setFont("Helvetica", self.config['fonts']['label_size'])
            c.drawString(x + padding, field_y - 2, "Contados:")
            pdfform.textFieldRelative(c, f"counted_{id(item)}",
                                     x + padding + 55, field_y - 15,
                                     width - 2*padding - 55, 15)

            # Boxes field
            field_y -= 20
            c.drawString(x + padding, field_y - 2, "Cajas:")
            pdfform.textFieldRelative(c, f"boxes_{id(item)}",
                                     x + padding + 55, field_y - 15,
                                     width - 2*padding - 55, 15)

            # Notes field
            field_y -= 20
            c.drawString(x + padding, field_y - 2, "Notas:")
            pdfform.textFieldRelative(c, f"notes_{id(item)}",
                                     x + padding + 55, field_y - 15,
                                     width - 2*padding - 55, 15)

    def _draw_image(self, c, image_path, x, y, cell_width, cell_height, padding):
        """Draw image centered in cell"""
        try:
            # Open image to get dimensions
            img = PILImage.open(image_path)
            img_width, img_height = img.size

            # Calculate scaling to fit within max dimensions
            max_width = self.config['image']['max_width']
            max_height = self.config['image']['max_height']

            scale = min(max_width / img_width, max_height / img_height, 1.0)
            display_width = img_width * scale
            display_height = img_height * scale

            # Center image in cell
            image_x = x + (cell_width - display_width) / 2
            image_y = y - padding - 30 - display_height  # Below type label

            # Draw image
            c.drawImage(image_path, image_x, image_y,
                       width=display_width, height=display_height,
                       preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"Warning: Could not load image {image_path}: {e}")
            # Draw placeholder text instead
            c.setFont("Helvetica", 8)
            c.drawCentredString(x + cell_width/2, y - cell_height/2, "[Image not found]")


def main():
    """Main entry point"""
    print("=" * 60)
    print("PDF Reference Sheet Generator")
    print("=" * 60)

    if len(sys.argv) < 2:
        print("\nUsage:")
        print("  python generate_reference.py <excel_file>")
        print("\nExample:")
        print("  python generate_reference.py order_data.xlsx")
        print("")
        sys.exit(1)

    excel_file = sys.argv[1]

    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)

    # Initialize generator
    generator = ReferenceSheetGenerator()

    # Read order data
    print(f"\nReading order data from: {excel_file}")
    order_data = generator.read_order_from_excel(excel_file)

    print(f"Order Number: {order_data['order_number']}")
    print(f"Client: {order_data['client_name']}")
    print(f"Items: {len(order_data['items'])}")

    # Generate PDF
    print("\nGenerating PDF...")
    output_file = generator.generate_pdf(order_data)

    print(f"\n[SUCCESS] Done! PDF saved to:")
    print(f"  {output_file}")
    print("\nYou can now:")
    print("  1. Open the PDF in Adobe Reader")
    print("  2. Fill in the editable fields (Contados, Cajas, Notas)")
    print("  3. Save the filled PDF")
    print("")


if __name__ == '__main__':
    main()
