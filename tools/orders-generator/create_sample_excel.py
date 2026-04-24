#!/usr/bin/env python3
"""Create sample Excel template for order data"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_sample_template():
    """Create sample Excel template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Data"

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="left", vertical="center")

    # Order information
    ws['A1'] = "Order Number"
    ws['B1'] = "001"
    ws['A2'] = "Client Name"
    ws['B2'] = "Sample Client"

    # Make order info bold
    ws['A1'].font = Font(bold=True)
    ws['A2'].font = Font(bold=True)

    # Column headers
    ws['A4'] = "Product Name"
    ws['B4'] = "Type"
    ws['C4'] = "Quantity"
    ws['D4'] = "Image Path"

    # Style headers
    for cell in ['A4', 'B4', 'C4', 'D4']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = header_alignment

    # Sample data
    # Using forward slashes (works on Windows too) or double backslashes
    sample_data = [
        ["Imán Ahualulco", "Imanes chicos", 50, "C:/images/iman_ahualulco.jpg"],
        ["Llavero Tequila", "Llaveros", 100, "C:/images/llavero_tequila.jpg"],
        ["Destapador Guadalajara", "Destapadores", 50, "C:/images/destapador_gdl.jpg"],
        ["Imán Chapala", "Imanes chicos", 75, "C:/images/iman_chapala.jpg"],
        ["Llavero Jalisco", "Llaveros", 80, "C:/images/llavero_jalisco.jpg"],
        ["Portallaves Tequila", "Portallaves", 60, "C:/images/portallaves.jpg"],
    ]

    for idx, row_data in enumerate(sample_data, start=5):
        ws[f'A{idx}'] = row_data[0]
        ws[f'B{idx}'] = row_data[1]
        ws[f'C{idx}'] = row_data[2]
        ws[f'D{idx}'] = row_data[3]

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50

    # Save workbook
    wb.save('sample_order.xlsx')
    print("Created sample_order.xlsx successfully!")

if __name__ == '__main__':
    create_sample_template()
