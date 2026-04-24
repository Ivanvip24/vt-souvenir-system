#!/usr/bin/env python3
"""
Quick PDF Reference Sheet Generator
Prompts for order name and generates PDF from sample_order.xlsx template
"""

import os
import tkinter as tk
from tkinter import simpledialog, messagebox
from datetime import datetime
from pathlib import Path

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase import pdfform
from PIL import Image as PILImage
from openpyxl import load_workbook
import yaml


class QuickPDFGenerator:
    """Quick PDF generator with dialog prompt"""

    def __init__(self):
        self.load_config()

    def load_config(self):
        """Load configuration"""
        with open('config.yaml', 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)

        # Set up page dimensions
        if self.config['layout']['page_size'] == 'A4':
            self.page_width, self.page_height = A4

        self.margin = self.config['layout']['margin'] * mm
        self.columns = self.config['layout']['columns']

    def ask_order_name(self):
        """Show dialog to ask for order name"""
        root = tk.Tk()
        root.withdraw()  # Hide main window

        # Center the dialog
        root.update_idletasks()

        order_name = simpledialog.askstring(
            "PDF Generator - Order Name",
            "Enter the order name/number:",
            parent=root
        )

        root.destroy()
        return order_name

    def ask_instructions(self):
        """Show dialog to ask for instructions"""
        root = tk.Tk()
        root.withdraw()  # Hide main window

        instructions = simpledialog.askstring(
            "PDF Generator - Instructions",
            "Enter instructions (optional):\n(e.g., '100 pzas of each design')",
            parent=root
        )

        root.destroy()
        return instructions if instructions else ""

    def ask_number_of_designs(self):
        """Show dialog with buttons to select number of designs"""
        root = tk.Tk()
        root.title("PDF Generator - Number of Designs")

        # Center the window
        window_width = 400
        window_height = 200
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        root.geometry(f"{window_width}x{window_height}+{x}+{y}")

        selected_value = tk.IntVar()
        selected_value.set(0)

        # Label
        label = tk.Label(root, text="How many designs does this order have?",
                        font=("Helvetica", 12))
        label.pack(pady=20)

        # Frame for buttons
        button_frame = tk.Frame(root)
        button_frame.pack(pady=10)

        def select_and_close(value):
            selected_value.set(value)
            root.quit()
            root.destroy()

        # Create buttons for common values
        common_values = [3, 5, 6]
        for i, value in enumerate(common_values):
            btn = tk.Button(button_frame, text=str(value),
                          command=lambda v=value: select_and_close(v),
                          width=10, height=2, font=("Helvetica", 14))
            btn.grid(row=0, column=i, padx=10)

        # Custom input option
        custom_frame = tk.Frame(root)
        custom_frame.pack(pady=10)

        custom_label = tk.Label(custom_frame, text="Other:", font=("Helvetica", 10))
        custom_label.pack(side=tk.LEFT, padx=5)

        custom_entry = tk.Entry(custom_frame, width=10, font=("Helvetica", 10))
        custom_entry.pack(side=tk.LEFT, padx=5)

        def custom_submit():
            try:
                value = int(custom_entry.get())
                if value > 0:
                    select_and_close(value)
            except ValueError:
                pass

        custom_btn = tk.Button(custom_frame, text="OK", command=custom_submit,
                             font=("Helvetica", 10))
        custom_btn.pack(side=tk.LEFT, padx=5)

        root.mainloop()
        return selected_value.get() if selected_value.get() > 0 else None

    def read_template_excel(self, excel_path='sample_order.xlsx'):
        """Read order data from Excel template"""
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Template file not found: {excel_path}")

        wb = load_workbook(excel_path)
        sheet = wb.active

        order_data = {
            'order_number': '',
            'client_name': '',
            'items': []
        }

        # Read items starting from row 5
        for row in range(5, sheet.max_row + 1):
            product_name = sheet.cell(row, 1).value
            product_type = sheet.cell(row, 2).value
            quantity = sheet.cell(row, 3).value
            image_path = sheet.cell(row, 4).value

            if not product_name:
                continue

            order_data['items'].append({
                'name': str(product_name),
                'type': str(product_type),
                'quantity': int(quantity) if quantity else 0,
                'image_path': str(image_path) if image_path else None
            })

        return order_data

    def generate_pdf(self, order_name, instructions, num_designs):
        """Generate PDF with the given order name, instructions, and number of designs"""

        # Create output directory
        output_dir = Path(self.config['output_path'])
        output_dir.mkdir(parents=True, exist_ok=True)

        # Create filename with order name
        output_filename = f"{order_name}.pdf"
        output_path = output_dir / output_filename

        # Create empty slots for designs
        empty_items = []
        for i in range(num_designs):
            empty_items.append({
                'name': f"Design {i+1}",
                'type': "",
                'quantity': 0,
                'image_path': None
            })

        # Create PDF
        c = canvas.Canvas(str(output_path), pagesize=(self.page_width, self.page_height))

        # Draw header with instructions
        self._draw_header(c, order_name, instructions, num_designs)

        # Draw grid with empty slots
        self._draw_items_grid(c, empty_items)

        # Save PDF
        c.save()

        return str(output_path)

    def _draw_header(self, c, order_name, instructions, num_designs):
        """Draw PDF header with instructions and logos"""
        # Navy blue and blue colors
        navy_blue = colors.HexColor('#1B4F72')  # Dark navy blue
        bright_blue = colors.HexColor('#2E86C1')  # Bright blue

        # Draw colored header background
        c.setFillColor(navy_blue)
        c.rect(0, self.page_height - 90, self.page_width, 90, fill=True, stroke=False)

        # Draw diagonal stripes accent (like reference)
        c.setFillColor(bright_blue)
        stripe_x = self.page_width - 150
        for i in range(6):
            x_offset = i * 15
            c.rect(stripe_x + x_offset, self.page_height - 90, 8, 90, fill=True, stroke=False)

        # Add VT logos
        logo_height = 60
        try:
            # Try to load vt.png (ANUNCIANDO text logo)
            if os.path.exists('vt.png'):
                c.drawImage('vt.png', self.margin, self.page_height - 75,
                           width=150, height=logo_height,
                           preserveAspectRatio=True, mask='auto')
        except:
            pass

        try:
            # Try to load vt1.png (VT symbol)
            if os.path.exists('vt1.png'):
                c.drawImage('vt1.png', self.page_width - self.margin - 80,
                           self.page_height - 75,
                           width=70, height=logo_height,
                           preserveAspectRatio=True, mask='auto')
        except:
            pass

        # Title
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(self.page_width / 2, self.page_height - 50, "REFERENCE SHEET")

        # Order info box (white background)
        y_pos = self.page_height - 100
        c.setFillColor(colors.white)
        c.setStrokeColor(navy_blue)
        c.setLineWidth(2)
        info_box_height = 25
        c.roundRect(self.margin, y_pos - info_box_height,
                   self.page_width - 2 * self.margin, info_box_height,
                   5, fill=True, stroke=True)

        # Order info text
        c.setFillColor(navy_blue)
        c.setFont("Helvetica-Bold", self.config['fonts']['body_size'])
        order_info = f"Order: {order_name} | Designs: {num_designs} | Date: {datetime.now().strftime('%Y-%m-%d')}"
        c.drawString(self.margin + 10, y_pos - 17, order_info)

        # Instructions field
        y_pos -= 35
        c.setFillColor(navy_blue)
        c.setFont("Helvetica-Bold", self.config['fonts']['body_size'])
        c.drawString(self.margin, y_pos, "Instructions:")

        # Draw instructions box with blue border
        c.setFont("Helvetica", self.config['fonts']['body_size'])
        box_x = self.margin + 80
        box_width = self.page_width - 2 * self.margin - 80
        box_height = 30

        # Colored border
        c.setStrokeColor(bright_blue)
        c.setFillColor(colors.HexColor('#EBF5FB'))  # Light blue background
        c.setLineWidth(2)
        c.roundRect(box_x, y_pos - box_height + 5, box_width, box_height, 3, fill=True, stroke=True)

        # Draw instructions text
        c.setFillColor(colors.black)
        if instructions:
            c.drawString(box_x + 5, y_pos - 10, instructions)

        y_pos -= box_height

        # CAJAS TOTALES field - very visible
        y_pos -= 40
        c.setFillColor(navy_blue)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(self.margin, y_pos + 5, "CAJAS TOTALES:")

        # Draw large box for total boxes with bright blue border
        cajas_box_x = self.margin + 120
        cajas_box_width = 150
        cajas_box_height = 35

        c.setStrokeColor(bright_blue)
        c.setFillColor(colors.HexColor('#EBF5FB'))
        c.setLineWidth(3)
        c.roundRect(cajas_box_x, y_pos - 15, cajas_box_width, cajas_box_height, 5, fill=True, stroke=True)

        # Add editable field for total boxes
        form = c.acroForm
        form.textfield(
            name="cajas_totales",
            x=cajas_box_x,
            y=y_pos - 15,
            width=cajas_box_width,
            height=cajas_box_height,
            borderWidth=0,
            textColor=colors.black,
            fontSize=16,
            fontName='Helvetica-Bold'
        )

        y_pos -= 25

        # Draw separator line with blue color
        c.setStrokeColor(bright_blue)
        c.setLineWidth(3)
        c.line(self.margin, y_pos - 5, self.page_width - self.margin, y_pos - 5)

    def _draw_items_grid(self, c, items):
        """Draw grid of items with dynamic sizing"""
        num_items = len(items)
        columns = 3  # Always 3 columns
        max_per_page = 9  # Maximum 3 rows x 3 columns

        # Calculate available space
        available_width = self.page_width - 2 * self.margin
        available_height = self.page_height - self.margin - 230  # Account for header with CAJAS TOTALES

        # Calculate number of items on first page
        items_on_page = min(num_items, max_per_page)
        rows_on_page = (items_on_page + columns - 1) // columns  # Ceiling division

        # Dynamic cell sizing with spacing
        spacing = 10  # Spacing between cells
        cell_width = (available_width - (columns - 1) * spacing) / columns
        cell_height = (available_height - (rows_on_page - 1) * spacing) / rows_on_page

        # Ensure minimum size
        cell_width = max(cell_width, 150)
        cell_height = max(cell_height, 180)

        padding = self.config['cell']['padding']

        # Center the grid
        total_grid_width = cell_width * columns + spacing * (columns - 1)
        total_grid_height = cell_height * rows_on_page + spacing * (rows_on_page - 1)

        start_x = (self.page_width - total_grid_width) / 2
        start_y = self.page_height - 230 - (available_height - total_grid_height) / 2

        current_x = start_x
        current_y = start_y
        items_in_row = 0
        items_drawn = 0

        for item in items:
            # Check if we need a new page
            if items_drawn > 0 and items_drawn % max_per_page == 0:
                c.showPage()

                # Recalculate for remaining items
                remaining_items = num_items - items_drawn
                items_on_page = min(remaining_items, max_per_page)
                rows_on_page = (items_on_page + columns - 1) // columns

                cell_height = (available_height - (rows_on_page - 1) * spacing) / rows_on_page
                cell_height = max(cell_height, 180)

                total_grid_height = cell_height * rows_on_page + spacing * (rows_on_page - 1)
                start_y = self.page_height - self.margin - (available_height - total_grid_height) / 2

                current_x = start_x
                current_y = start_y
                items_in_row = 0

            # Draw cell border
            c.setStrokeColor(colors.black)
            c.setLineWidth(1)
            c.rect(current_x, current_y - cell_height, cell_width, cell_height)

            # Draw item content
            self._draw_item_cell(c, item, current_x, current_y, cell_width, cell_height, padding)

            # Move to next position
            items_in_row += 1
            items_drawn += 1

            if items_in_row >= columns:
                current_y -= (cell_height + spacing)
                current_x = start_x
                items_in_row = 0
            else:
                current_x += (cell_width + spacing)

    def _draw_item_cell(self, c, item, x, y, width, height, padding):
        """Draw a single item cell"""
        navy_blue = colors.HexColor('#1B4F72')
        bright_blue = colors.HexColor('#2E86C1')
        light_blue = colors.HexColor('#EBF5FB')

        # Type label with blue background bar - make it editable
        c.setFillColor(bright_blue)
        c.rect(x, y - padding - 18, width, 18, fill=True, stroke=False)

        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(x + padding, y - padding - 12, "Tipo:")

        # Editable Type field
        type_field_x = x + padding + 30
        type_field_width = width - 2*padding - 30
        c.setFillColor(light_blue)
        c.rect(type_field_x, y - padding - 16, type_field_width, 14, fill=True, stroke=False)

        form = c.acroForm
        form.textfield(
            name=f"type_{id(item)}",
            x=type_field_x,
            y=y - padding - 16,
            width=type_field_width,
            height=14,
            borderWidth=0,
            textColor=colors.black,
            fontSize=9,
            fontName='Helvetica'
        )

        # Image area - centered placeholder
        image_y_center = y - padding - 30 - (height - 150) / 2
        c.setFillColor(colors.lightgrey)
        c.setFont("Helvetica", 10)
        c.drawCentredString(x + width/2, image_y_center, "[Image placeholder]")

        # Draw a light grey box for image area
        c.setStrokeColor(colors.grey)
        c.setLineWidth(0.5)
        image_box_height = height - 150
        c.rect(x + 10, y - padding - 30 - image_box_height, width - 20, image_box_height, stroke=True, fill=False)

        # Quantity ordered with blue accent - make it editable
        c.setFillColor(navy_blue)
        c.setFont("Helvetica-Bold", 9)
        qty_y = y - height + padding + 60
        c.drawString(x + padding, qty_y - 2, "Requeridos:")

        # Editable quantity field
        c.setFillColor(light_blue)
        c.setStrokeColor(bright_blue)
        c.setLineWidth(1)
        qty_field_x = x + padding + 55
        qty_field_width = width - 2*padding - 55
        c.rect(qty_field_x, qty_y - 15, qty_field_width, 15, fill=True, stroke=True)

        form = c.acroForm
        form.textfield(
            name=f"quantity_{id(item)}",
            x=qty_field_x,
            y=qty_y - 15,
            width=qty_field_width,
            height=15,
            borderWidth=0,
            textColor=colors.black,
            fontSize=10,
            fontName='Helvetica'
        )

        # Editable fields with blue labels
        if self.config['form_fields']['enable_editable']:
            field_y = qty_y - 25

            # Counted
            c.setFillColor(navy_blue)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(x + padding, field_y - 2, "Contados:")

            # Draw light blue background for field
            c.setFillColor(light_blue)
            c.setStrokeColor(bright_blue)
            c.setLineWidth(1)
            counted_field_x = x + padding + 55
            counted_field_width = width - 2*padding - 55
            c.rect(counted_field_x, field_y - 15, counted_field_width, 15, fill=True, stroke=True)

            # Add text field
            form = c.acroForm
            form.textfield(
                name=f"counted_{id(item)}",
                x=counted_field_x,
                y=field_y - 15,
                width=counted_field_width,
                height=15,
                borderWidth=0,
                textColor=colors.black,
                fontSize=10,
                fontName='Helvetica'
            )

    def _draw_image(self, c, image_path, x, y, cell_width, cell_height, padding):
        """Draw image centered and auto-sized in cell"""
        try:
            img = PILImage.open(image_path)
            img_width, img_height = img.size

            # Calculate available space for image (excluding header and footer areas)
            available_width = cell_width - 20  # 10px padding on each side
            available_height = cell_height - 150  # Space for tipo, requeridos, contados fields

            # Calculate scale to fit image within available space while maintaining aspect ratio
            scale_width = available_width / img_width
            scale_height = available_height / img_height
            scale = min(scale_width, scale_height)

            display_width = img_width * scale
            display_height = img_height * scale

            # Center the image horizontally and vertically in the available space
            image_x = x + (cell_width - display_width) / 2
            image_y = y - padding - 30 - available_height/2 - display_height/2

            c.drawImage(image_path, image_x, image_y,
                       width=display_width, height=display_height,
                       preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"Warning: Could not load image {image_path}: {e}")
            c.setFont("Helvetica", 10)
            c.setFillColor(colors.lightgrey)
            c.drawCentredString(x + cell_width/2, y - cell_height/2, "[Image not found]")


def main():
    """Main entry point"""
    print("=" * 60)
    print("Quick PDF Reference Sheet Generator")
    print("=" * 60)
    print()

    try:
        # Initialize generator
        generator = QuickPDFGenerator()

        # Ask for order name (shows dialog)
        print("Opening dialog box for order name...")
        order_name = generator.ask_order_name()

        if not order_name:
            print("No order name entered. Exiting.")
            return

        print(f"Order name: {order_name}")

        # Ask for instructions (shows second dialog)
        print("Opening dialog box for instructions...")
        instructions = generator.ask_instructions()

        if instructions:
            print(f"Instructions: {instructions}")
        else:
            print("No instructions provided")

        # Ask for number of designs (shows third dialog with buttons)
        print("Opening dialog box for number of designs...")
        num_designs = generator.ask_number_of_designs()

        if not num_designs or num_designs <= 0:
            print("No design count selected. Exiting.")
            return

        print(f"Number of designs: {num_designs}")

        # Generate PDF
        print("Generating PDF...")
        output_path = generator.generate_pdf(order_name, instructions, num_designs)

        print()
        print("=" * 60)
        print("[SUCCESS] PDF Generated!")
        print("=" * 60)
        print(f"File: {output_path}")
        print()
        print("You can now:")
        print("  1. Open the PDF in Adobe Reader")
        print("  2. Fill in the editable fields")
        print("  3. Save the completed PDF")
        print()

        # Show success dialog
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Success",
            f"PDF generated successfully!\n\nSaved to:\n{output_path}"
        )
        root.destroy()

    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        print("Please make sure sample_order.xlsx exists in the current folder.")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", str(e))
        root.destroy()

    except Exception as e:
        print(f"ERROR: {e}")
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{e}")
        root.destroy()


if __name__ == '__main__':
    main()
